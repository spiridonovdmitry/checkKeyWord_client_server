
import os

from chardet.universaldetector import UniversalDetector
import getpass
import docx2txt
from openpyxl import load_workbook
import time
import socket
from striprtf.striprtf import rtf_to_text
from docx import Document
import socket
import json
import fitz
import win32com.client

start_time = time.time()

def check_keywords_in_text(text, keywords):
    """""
    Функция check_keywords_in_text(text,keywords)
    Принимает аргументы text: Текст в котором происходит поиск.
                    keywords  - Список ключевых слов, которые 
                                мы проверяем на наличие
                    
    Что делает: Приводит текст к нижнему регистру, проверяет есть ли 
                в тексте ключевые слова из массива keywords. 
                    
    Возвращаемое значение: Возвращает массив ключевых слов, которые найдены в тексте. 
    """""
    found_keywords = [keyword for keyword in keywords if keyword.lower() in text.lower()]

    return found_keywords



def check_keywords_in_docx(file_path, keywords):
    """""
        функция считывает  
    """""
    content = ''
    text = ""
    try:
        content = Document(file_path)
        for paragraph in content.paragraphs:
            # Extract text from each paragraph
            text += paragraph.text + "\n"

    except Exception as e:
        k = 1
        print(f'Ошибка: {e} {file_path}')
    return check_keywords_in_text(text, keywords)


def check_keywords_in_doc(doc_path, keywords):
    """
    Extracts text from a DOC file and returns it as a string.

    :param doc_path: The path to the DOC file.
    :return: A string containing the extracted text.
    """
    try:
        # Create a new instance of the Word application
        word_app = win32com.client.Dispatch("Word.Application")

        # Open the DOC file
        doc = word_app.Documents.Open(os.path.abspath(doc_path))

        # Extract text from the document
        text = doc.Content.Text

        # Close the document
        doc.Close()

        # Quit the Word application
        word_app.Quit()

        return check_keywords_in_text(doc_path, keywords)
    except Exception as e:
        print(f"Error: {str(e)}")


def detection_encoding(file_path):
    """"
        Детектор кодировки. Не всегда работает((
        
        Принимает аргументы: file_path - путь для файла, кодировку 
                             которого мы хотим узнать
        Возвращает: Строку, значение которой является кодировка, если 
                    таковую удалось обнаружить
    """""
    try:
        detector = UniversalDetector()
        with open(file_path, 'r') as fh:
            for line in fh:
                detector.feed(line)
                if detector.done:
                    break
            detector.close()

    except Exception as e4:
        print(f'Ошибка: {e4} {file_path} {detector.result["encoding"]}')

    #Отсекает пустое значение, заменяя его кодировкой utf - 8
    if(detector.result["encoding"] == None ):
        return "utf-8"

    #print(detector.result["encoding"])
    return detector.result["encoding"]



def check_keywords_in_rtf(file_path, keywords):
    """""
        Функция, которая проверяет наличие ключевого слова в .rtf файле.
        
        Принимает аргументы: file_path - Путь до файла, который проверяем 
                             keywords  - Список ключевых слов, которые мы проверяем
                             на наличие.
        
        Что делает: Пытается считать файл по очереди различными кодировками.
                    В один момент доходит до детектора кодировки. 
                    Причём символы, кодировку которой установить не удалось, 
                    заменяются на ? знаки, что позволяет считать текст почти любой кодировкой.
                    Примечание: исходный текст никак не изменяется.       
                    
        Что возвращает: Список найденных ключевых слов.                           
    """""
    text = ''
    found_keywords = None
    try:
        with open(file_path) as infile:
            content = infile.read()
            text = rtf_to_text(content)
    except Exception as e:
        try:
            with open(file_path, encoding='cp1251') as infile:
                content = infile.read()
                text = rtf_to_text(content)
        except Exception as e2:
            try:
                with open(file_path, encoding='latin-1') as infile:
                    content = infile.read()
                    text = rtf_to_text(content)
            except Exception as e3:
                found_keywords = check_keywords_in_text(open(file_path, 'r', encoding=str(detection_encoding(file_path)), errors="ignore").read(), keywords)
                print(f'Ошибка: {e3} {file_path}')

    return found_keywords



def check_keywords_in_xlsx(file_path: str, keywords: list) -> str:
    """""
        Функция, которая проверяет наличие в файле ключевого слова
        
        Принимает аргументы: file_path - путь до файла, который проверяем.
                             keywords  - Список ключевых слов, наличие которых мы проверяем.
                             
        Что делает: Преобразует к тексту и посылает этот текст функции, которая работает в тексте.
                    
        Возвращает: Список найденных ключевых слов.                                 
    """""
    content = []
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    content.append(str(cell.value))
    except Exception as e:
        k = 1
      #  print(str('Ошибка:\n' + str(traceback.format_exc()) + ' ' + str(file_path) + ' КОД ' + str(e)))
    return check_keywords_in_text('\n'.join(content), keywords)



def check_keyword_in_txt(file_path, keywords):
    """""
        Функция, которая проверяет в файле 
    """""
    found_keywords = None
    try:
        found_keywords = check_keywords_in_text(open(file_path, 'r').read(), keywords)
    except Exception as e:
        #print(f'Ошибка: {e} {file_path}')
        try:
            found_keywords = check_keywords_in_text(open(file_path, 'r', encoding="UTF-16").read(), keywords)
        except Exception as e2:
            try:
                found_keywords = check_keywords_in_text(open(file_path, 'r', encoding="utf-8").read(), keywords)
            except Exception as e3:

                try:
                    detector = UniversalDetector()
                    with open(file_path, 'r', errors="ignore") as fh:
                        for line in fh:
                            detector.feed(line)
                            if detector.done:
                                break
                        detector.close()
                    found_keywords = check_keywords_in_text(open(file_path, 'r', errors="ignore", encoding=str(detector.result["encoding"])).read())
                except Exception as e4:
                    #print(f'Ошибка: {e3} {file_path} {detector.result["encoding"]}')
                    k = 1

    return found_keywords

def check_keywords_in_pdf(pdf_path, keywords):
    """
    Extracts text from a PDF file and returns it as a string.

    :param pdf_path: The path to the PDF file.
    :return: A string containing the extracted text.
    """
    try:
        # Open the PDF file
        pdf_document = fitz.open(pdf_path)

        # Initialize an empty string to store the text
        text = ""

        # Iterate through each page in the PDF
        for page_number in range(pdf_document.page_count):
            # Get the page
            page = pdf_document[page_number]

            # Extract text from the page
            page_text = page.get_text()

            # Append the extracted text to the overall text
            text += page_text

        # Close the PDF file
        pdf_document.close()

        return check_keywords_in_text(text, keywords)
    except Exception as e:
        # Handle any exceptions (e.g., file not found, invalid PDF, etc.)
        print( f"Error: {str(e)}")

def extension_processing(file_path, keywords):
    """
        Принимаемые аргументы: file - Путь до файла
                               Массив ключевых слов, которые нужно проверить в файле.

        Что делает: В зависимости от типа файла вызывает соответствующие обработчики.
                    Которые возвращают список найденных в файлах ключевых слов.

        Функция возвращает: список найденных в файле ключевых слов
    """
    found_keywords = None
    #if file_path.endswith(('.txt', '.docx', '.doc', '.rtf', '.xls', '.xlsx', '.pdf')):
    if file_path.endswith(('.txt')):
        if file_path.endswith('.txt'):
            found_keywords = check_keyword_in_txt(file_path, keywords)
        elif file_path.endswith(('.docx')):
            found_keywords = check_keywords_in_docx(file_path, keywords)
        elif file_path.endswith('.rtf'):
            found_keywords = check_keywords_in_rtf(file_path, keywords)
        elif file_path.endswith('.xlsx'):
            found_keywords = check_keywords_in_xlsx(file_path, keywords)
        elif file_path.endswith('.pdf'):
            found_keywords = check_keywords_in_pdf(file_path, keywords)
        elif file_path.endswith('.doc'):
            found_keywords = check_keywords_in_doc(file_path, keywords)
    return found_keywords



def search_files_in_folder(folder_path, keywords):
    """""
        Принимает аргументы: folder_path - изначальный путь, с которого начинается обход
                                keywords - массив ключевых слов, наличие которых мы хотим найти.
                                
        Что делает: Проходится по каждым папкам кроме архивов. Для каждого найденого файла вызывает 
                    функцию extension_processing и передаёт ей массив ключевых слов и путь до файла,
                    который программа пытается открыть. Если в файле были обнаружены ключевые слова,
                    то в логи записывается путь до файла и какое ключевое слово обнаружено.
                    
        Возвращает: Лог- то есть массив записей, первой записью которого будет имя пользователя.                                   
    """""

    name = os.environ['USERNAME']
    log = [name]
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:

            file_path = os.path.join(root, file_name)
            if file_path.endswith(('.zip', '.rar', '.7z')):
                continue
            else:
                found_keywords = extension_processing(file_path, keywords)
            if found_keywords:
                add_log(name, file_path, found_keywords, log)
    return log



def add_log(name, file_path, found_keywords, log):
    """""
        Добавляет в log новую запись, содержашую путь до файла и ключевое слово   
    """""
    log.append(f'{file_path} {found_keywords}')


def start_client(log_final):
    host = '127.0.0.1'
    port = 12345

    client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    client_socket.connect((host, port))

    json_data = json.dumps(log_final)
    client_socket.send(json_data.encode('utf-8'))

    #receive_file(client_socket, file_name)


if __name__ == "__main__":
    # Путь к папке, которую нужно проверить
    folder_path = 'C:\\'
    # Список ключевых слов для поиска
    keywords = ['пример', 'test']
    log = search_files_in_folder(folder_path, keywords)
    print([str(x) + ' ' for x in log])
    start_client(log)

end_time = time.time()
# Рассчитайте разницу, чтобы узнать время выполнения
execution_time = end_time - start_time
print(str(execution_time))

with open("./out.txt", "w") as file:
    #log = file.read()
    #start_client(log)
    for one_log in log:
        file.write(one_log+"\n")


######## сокеты
